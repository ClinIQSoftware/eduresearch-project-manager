from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
from io import BytesIO
import secrets
import string
from app.database import get_db
from app.models.user import User
from app.models.project import Project
from app.models.email_settings import EmailSettings
from app.models.email_template import EmailTemplate
from app.models.system_settings import SystemSettings
from app.models.organization import Institution
from app.models.department import Department
from app.schemas.user import UserResponse, UserUpdateAdmin, UserCreateAdmin, PendingUserResponse
from app.schemas.project import ProjectResponse, ProjectUpdate
from app.schemas.email_settings import (
    EmailSettingsCreate, EmailSettingsUpdate, EmailSettingsResponse
)
from app.schemas.email_template import (
    EmailTemplateResponse, EmailTemplateUpdate, TestEmailRequest
)
from app.schemas.system_settings import (
    SystemSettingsResponse, SystemSettingsUpdate, BulkUploadResult
)
from app.dependencies import get_current_user, require_superuser, is_institution_admin, is_institution_admin
from app.services.auth import create_user, get_password_hash
from app.services.email import email_service

router = APIRouter()


# User Management

@router.get("/users", response_model=List[UserResponse])
def get_all_users(
    institution_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all users (admin only)."""
    query = db.query(User)

    if current_user.is_superuser:
        if institution_id:
            query = query.filter(User.institution_id == institution_id)
    else:
        # Institution admin can only see their institution's users
        if current_user.institution_id:
            if not is_institution_admin(db, current_user.id, current_user.institution_id):
                raise HTTPException(status_code=403, detail="Admin access required")
            query = query.filter(User.institution_id == current_user.institution_id)
        else:
            raise HTTPException(status_code=403, detail="Admin access required")

    return query.order_by(User.last_name, User.first_name).all()


@router.post("/users", response_model=UserResponse)
async def create_user_admin(
    user_data: UserCreateAdmin,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new user (admin only). Auto-generates password and sends welcome email."""
    # Check admin access
    if not current_user.is_superuser:
        if user_data.institution_id:
            if not is_institution_admin(db, current_user.id, user_data.institution_id):
                raise HTTPException(status_code=403, detail="Admin access required")
        else:
            raise HTTPException(status_code=403, detail="Superuser access required")

    # Only superuser can create superusers
    if user_data.is_superuser and not current_user.is_superuser:
        raise HTTPException(status_code=403, detail="Only superuser can create superuser accounts")

    # Check if email exists
    existing = db.query(User).filter(User.email == user_data.email).first()
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    # Generate temporary password
    temp_password = generate_temp_password()

    user = create_user(
        db=db,
        email=user_data.email,
        password=temp_password,
        first_name=user_data.first_name,
        last_name=user_data.last_name,
        institution_id=user_data.institution_id,
        department_id=user_data.department_id,
        is_superuser=user_data.is_superuser
    )

    # Send welcome email with temporary password in background
    full_name = f"{user_data.first_name} {user_data.last_name}".strip()
    background_tasks.add_task(
        email_service.send_welcome_email,
        user_data.email,
        full_name,
        temp_password
    )

    return user


@router.put("/users/{user_id}", response_model=UserResponse)
def update_user_admin(
    user_id: int,
    user_data: UserUpdateAdmin,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update a user (admin only)."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Check admin access
    if not current_user.is_superuser:
        if user.institution_id:
            if not is_institution_admin(db, current_user.id, user.institution_id):
                raise HTTPException(status_code=403, detail="Admin access required")
        else:
            raise HTTPException(status_code=403, detail="Superuser access required")

    # Only superuser can change superuser status
    if user_data.is_superuser is not None and not current_user.is_superuser:
        raise HTTPException(status_code=403, detail="Only superuser can grant superuser status")

    update_data = user_data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(user, key, value)

    db.commit()
    db.refresh(user)
    return user


@router.delete("/users/{user_id}")
def deactivate_user(
    user_id: int,
    current_user: User = Depends(require_superuser),
    db: Session = Depends(get_db)
):
    """Deactivate a user (superuser only)."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot deactivate yourself")

    user.is_active = False
    db.commit()

    return {"message": "User deactivated"}


@router.delete("/users/{user_id}/permanent")
def delete_user_permanently(
    user_id: int,
    current_user: User = Depends(require_superuser),
    db: Session = Depends(get_db)
):
    """Permanently delete a user (superuser only). This action cannot be undone."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")

    if user.is_superuser:
        # Count other superusers
        other_superusers = db.query(User).filter(
            User.is_superuser == True,
            User.id != user_id
        ).count()
        if other_superusers == 0:
            raise HTTPException(
                status_code=400,
                detail="Cannot delete the last superuser"
            )

    db.delete(user)
    db.commit()

    return {"message": "User permanently deleted"}


# Project Management (Superuser)

@router.get("/projects", response_model=List[ProjectResponse])
def get_all_projects_admin(
    current_user: User = Depends(require_superuser),
    db: Session = Depends(get_db)
):
    """Get all projects (superuser only)."""
    return db.query(Project).order_by(Project.created_at.desc()).all()


@router.put("/projects/{project_id}", response_model=ProjectResponse)
def update_project_admin(
    project_id: int,
    project_data: ProjectUpdate,
    current_user: User = Depends(require_superuser),
    db: Session = Depends(get_db)
):
    """Update any project (superuser only)."""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    update_data = project_data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(project, key, value)

    db.commit()
    db.refresh(project)
    return project


# Email Settings

@router.get("/email-settings", response_model=EmailSettingsResponse)
def get_email_settings(
    institution_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get email settings."""
    # Check admin access
    if not current_user.is_superuser:
        if institution_id:
            if not is_institution_admin(db, current_user.id, institution_id):
                raise HTTPException(status_code=403, detail="Admin access required")
        else:
            raise HTTPException(status_code=403, detail="Superuser access required")

    settings = db.query(EmailSettings).filter(
        EmailSettings.institution_id == institution_id
    ).first()

    if not settings:
        # Return default settings
        return EmailSettingsResponse(
            id=0,
            institution_id=institution_id,
            smtp_host="smtp.gmail.com",
            smtp_port=587,
            from_name="EduResearch Project Manager",
            is_active=False
        )

    return settings


@router.put("/email-settings", response_model=EmailSettingsResponse)
def update_email_settings(
    settings_data: EmailSettingsUpdate,
    institution_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update email settings."""
    # Check admin access
    if not current_user.is_superuser:
        if institution_id:
            if not is_institution_admin(db, current_user.id, institution_id):
                raise HTTPException(status_code=403, detail="Admin access required")
        else:
            raise HTTPException(status_code=403, detail="Superuser access required")

    settings = db.query(EmailSettings).filter(
        EmailSettings.institution_id == institution_id
    ).first()

    if not settings:
        # Create new settings
        settings = EmailSettings(institution_id=institution_id)
        db.add(settings)

    update_data = settings_data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(settings, key, value)

    db.commit()
    db.refresh(settings)
    return settings


# System Settings

@router.get("/system-settings", response_model=SystemSettingsResponse)
def get_system_settings(
    institution_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get system settings (superuser only for global, org admin for org-specific)."""
    if not current_user.is_superuser:
        if institution_id:
            if not is_institution_admin(db, current_user.id, institution_id):
                raise HTTPException(status_code=403, detail="Admin access required")
        else:
            raise HTTPException(status_code=403, detail="Superuser access required")

    settings = db.query(SystemSettings).filter(
        SystemSettings.institution_id == institution_id
    ).first()

    if not settings:
        # Return default settings
        return SystemSettingsResponse(
            id=0,
            institution_id=institution_id,
            require_registration_approval=False,
            registration_approval_mode="block",
            min_password_length=8,
            require_uppercase=True,
            require_lowercase=True,
            require_numbers=True,
            require_special_chars=False,
            session_timeout_minutes=30,
            google_oauth_enabled=True,
            microsoft_oauth_enabled=True
        )

    return settings


@router.put("/system-settings", response_model=SystemSettingsResponse)
def update_system_settings(
    settings_data: SystemSettingsUpdate,
    institution_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update system settings (superuser only for global, org admin for org-specific)."""
    if not current_user.is_superuser:
        if institution_id:
            if not is_institution_admin(db, current_user.id, institution_id):
                raise HTTPException(status_code=403, detail="Admin access required")
        else:
            raise HTTPException(status_code=403, detail="Superuser access required")

    settings = db.query(SystemSettings).filter(
        SystemSettings.institution_id == institution_id
    ).first()

    if not settings:
        # Create new settings
        settings = SystemSettings(institution_id=institution_id)
        db.add(settings)

    update_data = settings_data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(settings, key, value)

    db.commit()
    db.refresh(settings)
    return settings


# User Approval

@router.get("/pending-users", response_model=List[PendingUserResponse])
def get_pending_users(
    institution_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get users pending approval."""
    if not current_user.is_superuser:
        if institution_id:
            if not is_institution_admin(db, current_user.id, institution_id):
                raise HTTPException(status_code=403, detail="Admin access required")
        else:
            raise HTTPException(status_code=403, detail="Superuser access required")

    query = db.query(User).filter(User.is_approved == False)

    if institution_id:
        query = query.filter(User.institution_id == institution_id)
    elif not current_user.is_superuser and current_user.institution_id:
        query = query.filter(User.institution_id == current_user.institution_id)

    return query.order_by(User.created_at.desc()).all()


@router.post("/approve-user/{user_id}", response_model=UserResponse)
def approve_user(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Approve a pending user registration."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    if user.is_approved:
        raise HTTPException(status_code=400, detail="User is already approved")

    # Check admin access
    if not current_user.is_superuser:
        if user.institution_id:
            if not is_institution_admin(db, current_user.id, user.institution_id):
                raise HTTPException(status_code=403, detail="Admin access required")
        else:
            raise HTTPException(status_code=403, detail="Superuser access required")

    user.is_approved = True
    user.approved_at = datetime.utcnow()
    user.approved_by_id = current_user.id

    db.commit()
    db.refresh(user)
    return user


@router.post("/reject-user/{user_id}")
def reject_user(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Reject and delete a pending user registration."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    if user.is_approved:
        raise HTTPException(status_code=400, detail="Cannot reject an approved user")

    # Check admin access
    if not current_user.is_superuser:
        if user.institution_id:
            if not is_institution_admin(db, current_user.id, user.institution_id):
                raise HTTPException(status_code=403, detail="Admin access required")
        else:
            raise HTTPException(status_code=403, detail="Superuser access required")

    db.delete(user)
    db.commit()

    return {"message": "User registration rejected and deleted"}


# Bulk User Upload

def generate_temp_password(length: int = 12) -> str:
    """Generate a secure temporary password."""
    alphabet = string.ascii_letters + string.digits + "!@#$%"
    return ''.join(secrets.choice(alphabet) for _ in range(length))


@router.post("/users/bulk-upload", response_model=BulkUploadResult)
async def bulk_upload_users(
    file: UploadFile = File(...),
    current_user: User = Depends(require_superuser),
    db: Session = Depends(get_db)
):
    """
    Bulk upload users from Excel file.
    Expected columns: email, first_name, last_name, phone, bio, institution, department, is_superuser
    Institution and department can be specified by name or by ID (institution_id, department_id).
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")

    try:
        import openpyxl

        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents))
        sheet = workbook.active

        # Get headers from first row
        headers = [cell.value.lower().strip() if cell.value else '' for cell in sheet[1]]

        required_columns = ['email', 'first_name', 'last_name']
        for col in required_columns:
            if col not in headers:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing required column: {col}"
                )

        # Pre-load institutions and departments for name lookup
        institutions_by_name = {i.name.lower(): i for i in db.query(Institution).all()}
        departments_by_name = {}
        for d in db.query(Department).all():
            key = (d.institution_id, d.name.lower())
            departments_by_name[key] = d

        created = 0
        skipped = 0
        errors = []

        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue

            row_data = dict(zip(headers, row))

            email = row_data.get('email', '').strip() if row_data.get('email') else ''
            first_name = row_data.get('first_name', '').strip() if row_data.get('first_name') else ''
            last_name = row_data.get('last_name', '').strip() if row_data.get('last_name') else ''

            if not email or not first_name or not last_name:
                errors.append(f"Row {row_num}: Missing email, first_name, or last_name")
                continue

            # Check if user already exists
            existing = db.query(User).filter(User.email == email).first()
            if existing:
                skipped += 1
                continue

            try:
                # Generate temporary password
                temp_password = generate_temp_password()

                # Get optional fields
                phone = row_data.get('phone', '').strip() if row_data.get('phone') else None
                bio = row_data.get('bio', '').strip() if row_data.get('bio') else None

                # Handle institution - by name or by ID
                inst_id = None
                inst_name = row_data.get('institution', '').strip() if row_data.get('institution') else ''
                inst_id_val = row_data.get('institution_id')

                if inst_name:
                    # Lookup by name
                    inst = institutions_by_name.get(inst_name.lower())
                    if inst:
                        inst_id = inst.id
                    else:
                        errors.append(f"Row {row_num}: Institution '{inst_name}' not found")
                        continue
                elif inst_id_val and str(inst_id_val).strip():
                    # Use ID directly
                    try:
                        inst_id = int(inst_id_val)
                    except (ValueError, TypeError):
                        inst_id = None

                # Handle department - by name or by ID
                dept_id = None
                dept_name = row_data.get('department', '').strip() if row_data.get('department') else ''
                dept_id_val = row_data.get('department_id')

                if dept_name:
                    # Lookup by name (requires institution_id)
                    if inst_id:
                        dept_key = (inst_id, dept_name.lower())
                        dept = departments_by_name.get(dept_key)
                        if dept:
                            dept_id = dept.id
                        else:
                            errors.append(f"Row {row_num}: Department '{dept_name}' not found in the specified institution")
                            continue
                    else:
                        errors.append(f"Row {row_num}: Department specified but no institution provided")
                        continue
                elif dept_id_val and str(dept_id_val).strip():
                    # Use ID directly
                    try:
                        dept_id = int(dept_id_val)
                    except (ValueError, TypeError):
                        dept_id = None

                is_superuser = False
                superuser_val = row_data.get('is_superuser', '')
                if superuser_val:
                    is_superuser = str(superuser_val).lower() in ('true', 'yes', '1', 'y')

                # Create user
                user = User(
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    password_hash=get_password_hash(temp_password),
                    phone=phone,
                    bio=bio,
                    institution_id=inst_id,
                    department_id=dept_id,
                    is_superuser=is_superuser,
                    is_approved=True,  # Auto-approve bulk uploaded users
                    is_active=True
                )
                db.add(user)
                created += 1

                # TODO: Send email with temporary password

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.commit()

        return BulkUploadResult(
            total_processed=created + skipped + len(errors),
            created=created,
            skipped=skipped,
            errors=errors
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@router.get("/users/upload-template")
async def get_upload_template(
    current_user: User = Depends(require_superuser)
):
    """Download Excel template for bulk user upload."""
    import openpyxl
    from fastapi.responses import StreamingResponse

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Users"

    # Add headers - include both name and ID columns for flexibility
    headers = ['email', 'first_name', 'last_name', 'phone', 'bio', 'institution', 'department', 'is_superuser']
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)

    # Add example row
    example = ['user@example.com', 'John', 'Doe', '+1234567890', 'Bio text', 'Example University', 'Research Department', 'false']
    for col, value in enumerate(example, start=1):
        sheet.cell(row=2, column=col, value=value)

    # Add instructions row
    instructions = ['Required', 'Required', 'Required', 'Optional', 'Optional', 'Optional - use institution name', 'Optional - use department name', 'Optional - true/false']
    for col, value in enumerate(instructions, start=1):
        cell = sheet.cell(row=3, column=col, value=value)
        cell.font = openpyxl.styles.Font(italic=True, color="808080")

    # Save to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=user_upload_template.xlsx"}
    )


# Email Templates

@router.get("/email-templates", response_model=List[EmailTemplateResponse])
def get_email_templates(
    institution_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all email templates."""
    # Check admin access
    if not current_user.is_superuser:
        if institution_id:
            if not is_institution_admin(db, current_user.id, institution_id):
                raise HTTPException(status_code=403, detail="Admin access required")
        else:
            raise HTTPException(status_code=403, detail="Superuser access required")

    # Get institution-specific templates first, then fall back to global
    templates = db.query(EmailTemplate).filter(
        EmailTemplate.institution_id == institution_id
    ).all()

    # If no institution-specific templates, get global ones
    if not templates:
        templates = db.query(EmailTemplate).filter(
            EmailTemplate.institution_id.is_(None)
        ).all()

    return templates


@router.get("/email-templates/{template_type}", response_model=EmailTemplateResponse)
def get_email_template(
    template_type: str,
    institution_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get a specific email template by type."""
    # Check admin access
    if not current_user.is_superuser:
        if institution_id:
            if not is_institution_admin(db, current_user.id, institution_id):
                raise HTTPException(status_code=403, detail="Admin access required")
        else:
            raise HTTPException(status_code=403, detail="Superuser access required")

    # Try institution-specific first
    template = db.query(EmailTemplate).filter(
        EmailTemplate.template_type == template_type,
        EmailTemplate.institution_id == institution_id
    ).first()

    # Fall back to global
    if not template:
        template = db.query(EmailTemplate).filter(
            EmailTemplate.template_type == template_type,
            EmailTemplate.institution_id.is_(None)
        ).first()

    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    return template


@router.put("/email-templates/{template_type}", response_model=EmailTemplateResponse)
def update_email_template(
    template_type: str,
    template_data: EmailTemplateUpdate,
    institution_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update an email template."""
    # Check admin access
    if not current_user.is_superuser:
        if institution_id:
            if not is_institution_admin(db, current_user.id, institution_id):
                raise HTTPException(status_code=403, detail="Admin access required")
        else:
            raise HTTPException(status_code=403, detail="Superuser access required")

    # Try to get existing template for this institution
    template = db.query(EmailTemplate).filter(
        EmailTemplate.template_type == template_type,
        EmailTemplate.institution_id == institution_id
    ).first()

    if not template:
        # Get the global template to copy as base
        global_template = db.query(EmailTemplate).filter(
            EmailTemplate.template_type == template_type,
            EmailTemplate.institution_id.is_(None)
        ).first()

        if not global_template:
            raise HTTPException(status_code=404, detail="Template type not found")

        # Create institution-specific template
        template = EmailTemplate(
            institution_id=institution_id,
            template_type=template_type,
            subject=global_template.subject,
            body=global_template.body,
            is_active=global_template.is_active
        )
        db.add(template)

    # Update with provided data
    update_data = template_data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(template, key, value)

    db.commit()
    db.refresh(template)
    return template


@router.post("/email-templates/test")
async def send_test_email(
    request: TestEmailRequest,
    background_tasks: BackgroundTasks,
    institution_id: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Send a test email using a specific template."""
    # Check admin access
    if not current_user.is_superuser:
        if institution_id:
            if not is_institution_admin(db, current_user.id, institution_id):
                raise HTTPException(status_code=403, detail="Admin access required")
        else:
            raise HTTPException(status_code=403, detail="Superuser access required")

    # Get the template
    template = db.query(EmailTemplate).filter(
        EmailTemplate.template_type == request.template_type,
        EmailTemplate.institution_id == institution_id
    ).first()

    if not template:
        template = db.query(EmailTemplate).filter(
            EmailTemplate.template_type == request.template_type,
            EmailTemplate.institution_id.is_(None)
        ).first()

    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    # Sample context for test
    test_context = {
        "user_name": "Test User",
        "user_email": request.recipient_email,
        "institution_name": "Test Institution",
        "department_name": "Test Department",
        "approval_link": "https://example.com/admin/pending-users",
        "project_name": "Test Project",
        "requester_name": "Test Requester",
        "message": "This is a test join request message.",
        "project_link": "https://example.com/projects/1",
        "task_title": "Test Task",
        "priority": "High",
        "due_date": "2025-01-15",
        "description": "This is a test task description.",
        "task_link": "https://example.com/tasks/1"
    }

    background_tasks.add_task(
        email_service.send_templated_email,
        request.recipient_email,
        template.subject,
        template.body,
        test_context
    )

    return {"message": f"Test email queued for {request.recipient_email}"}
